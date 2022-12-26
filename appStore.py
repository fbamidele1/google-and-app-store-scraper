import random
import time
from openpyxl import load_workbook
from pprint import pprint
from app_store_scraper import AppStore
from appid import get_app_id
"""
iTunes App Store Scraper
"""
import requests
import json
import time
import re
import os
from datetime import datetime

from urllib.parse import quote_plus
from itunes_app_scraper.util import AppStoreException, AppStoreCollections, AppStoreMarkets, COUNTRIES

class Regex:
	STARS = re.compile(r"<span class=\"total\">[\s\S]*?</span>")

term = 'snapchat'
class AppStoreScraper:
	


	def get_app_details(self, app_id, country="us", lang="", add_ratings=False, flatten=True, sleep=None, force=False):
		
		try:
			app_id = int(app_id)
			id_field = "id"
		except ValueError:
			id_field = "bundleId"

		if force:
			# this will by-pass the serverside caching
			import secrets
			timestamp = secrets.token_urlsafe(8)
			url = "https://itunes.apple.com/lookup?%s=%s&country=%s&entity=software&timestamp=%s" % (id_field, app_id, country, timestamp)
		else:
			url = "https://itunes.apple.com/lookup?%s=%s&country=%s&entity=software" % (id_field, app_id, country)

		try:
			if sleep is not None:
				time.sleep(sleep)
			result = requests.get(url).json()
		except Exception:
			try:
				# handle the retry here.
				# Take an extra sleep as back off and then retry the URL once.
				time.sleep(2)
				result = requests.get(url).json()
			except Exception:
				raise AppStoreException("Could not parse app store response for ID %s" % app_id)

		try:
			app = result["results"][0]
		except (KeyError, IndexError):
			raise AppStoreException("No app found with ID %s" % app_id)

		if add_ratings:
			try:
				ratings = self.get_app_ratings(app_id, countries=country)
				app['user_ratings'] = ratings
			except AppStoreException:
				# Return some details
				self._log_error(country, 'Unable to collect ratings for %s' % str(app_id))
				app['user_ratings'] = 'Error; unable to collect ratings'

		# 'flatten' app response
		# responses are at most two-dimensional (array within array), so simply
		# join any such values
		if flatten:
			for field in app:
				if isinstance(app[field], list):
					app[field] = ",".join(app[field])
				elif isinstance(app[field], dict):
					app[field] = ", ".join(["%s star: %s" % (key, value) for key,value in app[field].items()])

		return app


	def get_store_id_for_country(self, country):
		country = country.upper()

		if hasattr(AppStoreMarkets, country):
			return getattr(AppStoreMarkets, country)
		else:
			raise AppStoreException("Country code not found for {0}".format(country))

	def get_app_ratings(self, app_id, countries=['us', 'nl'], sleep=1):
		dataset = { 1: 0, 2: 0, 3: 0, 4: 0, 5: 0 }
		if countries is None:
			countries = COUNTRIES
		elif isinstance(countries, str): # only a string provided
			countries = [countries]
		else:
			countries = countries

		for country in countries:
			url = "https://itunes.apple.com/%s/customer-reviews/id%s?displayable-kind=11" % (country, app_id)
			store_id = self.get_store_id_for_country(country)
			headers = { 'X-Apple-Store-Front': '%s,12 t:native' % store_id }

			try:
				if sleep is not None:
					time.sleep(sleep)
				result = requests.get(url, headers=headers).text
			except Exception:
				try:
					# handle the retry here.
					# Take an extra sleep as back off and then retry the URL once.
					time.sleep(2)
					result = requests.get(url, headers=headers).text
				except Exception:
					raise AppStoreException("Could not parse app store rating response for ID %s" % app_id)

			ratings = self._parse_rating(result)

			if ratings is not None:
				dataset[1] = dataset[1] + ratings[1]
				dataset[2] = dataset[2] + ratings[2]
				dataset[3] = dataset[3] + ratings[3]
				dataset[4] = dataset[4] + ratings[4]
				dataset[5] = dataset[5] + ratings[5]

        # debug
		#,print("-----------------------")
		#,print('%d ratings' % (dataset[1] + dataset[2] + dataset[3] + dataset[4] + dataset[5]))
		#,print(dataset)

		return dataset

	def _parse_rating(self, text):
		matches = Regex.STARS.findall(text)

		if len(matches) != 5:
			# raise AppStoreException("Cant get stars - expected 5 - but got %d" % len(matches))
			return None

		ratings = {}
		star = 5

		for match in matches:
			value = match
			value = value.replace("<span class=\"total\">", "")
			value = value.replace("</span>", "")
			ratings[star] = int(value)
			star = star - 1

		return ratings

	def _log_error(self, app_store_country, message):
		"""
		Write the error to a local file to capture the error.
		:param str app_store_country: the country for the app store
		:param str message: the error message to log
		"""
		log_dir = 'log/'
		if not os.path.isdir(log_dir):
			os.mkdir(log_dir)

		app_log = os.path.join(log_dir, "{0}_log.txt".format(app_store_country))
		errortime = datetime.now().strftime('%Y%m%d_%H:%M:%S - ')
		fh = open(app_log, "a")
		fh.write("%s %s \n" % (errortime,message))
		fh.close()

def details_reviews(app):
	print('############# Data scraping started...|')
    # Excel#########################################
	wb = load_workbook('AppData.xlsx')
	ws = wb['Details']
    # Excel#########################################
	name = app['trackName']
	app_id = app['trackId']
	#  For review count
	app_rev = AppStore(country="us", app_name=name, app_id = app_id )
	app_rev.review(how_many=100)
	#  For review count
	publisher = app['sellerName']
	icon_url = app['artworkUrl60']
	publisher_email = ''
	try:
		website = app['sellerUrl']
	except:
		website = ''
	category = app['primaryGenreName']
	avg_rating = app['averageUserRating']
	rating_count = app['userRatingCount']
	# 5 star missing
	# 5 reviews_count missing
	release_date = app['releaseDate']
	from dateutil import parser
	new_date = parser.isoparse(release_date)
	new_release_date = new_date.strftime('%m/%d/%Y  %H:%M:%S')
	updated = app['currentVersionReleaseDate']
	new_update = parser.isoparse(updated)
	new_updated = new_update.strftime('%m/%d/%Y  %H:%M:%S')
	description = app['description']
	version = app['version']
	free = app['formattedPrice']
	if free == 'Free':
		free = True
	else:
		free = False
	
	developer_id = app['artistId']
	app_url = app['trackViewUrl']
	store = 'App Store'
	
	app_d = AppStoreScraper()
	rating = app_d.get_app_ratings(app_id)
	star_5 = rating[5]
	star_4 = rating[4]
	star_3 = rating[3]
	star_2 = rating[2]
	star_1 = rating[1]
	review_coun = app_rev.reviews_count
	app_data = [name, icon_url, publisher, publisher_email, website, category, avg_rating, rating_count, star_5, star_4, star_3, star_2, star_1, review_coun, new_release_date, new_updated, description, version, free, developer_id, app_url, app_id, store]
	# print(app_data)
	ws.append(app_data)
	wb.save('AppData.xlsx')
	# Fetching Reviews
	reviews = app_rev.reviews
	wait_time = random.randint(1,5)
	print(f'############# {review_coun} reviews fount for {app_id} ID! #############')
	for review in reviews:
	# Excel#########################################
		wb = load_workbook('AppData.xlsx')
		ws = wb['Reviews']
		# Excel#########################################
		print('############# Saving review data #############')
		date = review['date']
		review_date = date.strftime('%m/%d/%Y  %H:%M:%S')
		author_name = review['userName']
		review_title = review['title']
		review_content = review['review']
		star = review['rating']
		store = 'App Store'
		version = ''
		review_data = [name, app_id, publisher, review_date, author_name, review_title, review_content, star, version, store]
		ws.append(review_data)
		wb.save('AppData.xlsx')
		print(review_data)
		print('############# Review data successfully saved! #############')
		time.sleep(wait_time)
	wb.close()



def main():
	wait_time = random.randint(1,5)
	app_d = AppStoreScraper()
	app_ids = get_app_id()

	for app_id in app_ids:
		app = app_d.get_app_details(app_id)
		details_reviews(app)
		time.sleep(wait_time)


if __name__ == '__main__':
	main()