from google_play_scraper import app, reviews, Sort
from openpyxl import load_workbook
from appid import get_play_id
from pprint import pprint

def details(appid):
    wb = load_workbook('AppData.xlsx')
    ws = wb['Details']
    result = app(
        appid,
        lang='en', # defaults to 'en'
        country='us' # defaults to 'us'
    )
    pprint(result)
    name = result['title']
    icon_url = result['icon']
    publisher = result['developer']
    publisher_email = result['developerEmail']
    website = result['developerWebsite']
    category = result['genreId']
    rating = result['score']
    rating_count = result['ratings']
    star_5 = result['histogram'][-1]
    star_4 = result['histogram'][-2]
    star_3 = result['histogram'][-3]
    star_2 = result['histogram'][1]
    star_1 = result['histogram'][0]
    review_count = result['reviews']
    release_date = result['released']
    updated = result['updated']
    description = result['description']
    version = result['version']
    free = result['free']
    developer_id = result['developerId']
    app_url = result['url']
    app_id = result['appId']
    store = 'Play Store'
    
    app_data = [name, icon_url, publisher, publisher_email, website, category, rating, rating_count, star_5, star_4, star_3, star_2, star_1, review_count, release_date, updated, description, version, free, developer_id, app_url, app_id, store]
    ws.append(app_data)
    wb.save('AppData.xlsx')
    wb.close()
    print(f'{appid}: Details database created successfully!')
    return app_data

def review_all(appid):
    wb = load_workbook('AppData.xlsx')
    ws = wb['Reviews']
    result, continuation_token = reviews(
    appid,
    lang='en', # defaults to 'en'
    country='us', # defaults to 'us'
    sort=Sort.NEWEST, # defaults to Sort.NEWEST
    count= 100, # defaults to 100
    #filter_score_with=5 # defaults to None(means all score)
    )
    detail = app(
        appid,
        lang='en', # defaults to 'en'
        country='us' # defaults to 'us'
    )
    name = detail['title']
    application_id = detail['appId']
    application_name = detail['developer']
    
    
    for r in result:
        data = [
        name,
        application_id,
        application_name
    ]
        date_time = r['at']
        review_date = date_time.strftime('%m/%d/%Y  %H:%M:%S')
        author_name = r['userName']
        review_title = ''
        review_content = r['content']
        stars = r['score']
        version = r['reviewCreatedVersion']
        store = 'Play Store'
        review_final_data = [data[0],data[1], data[2], review_date, author_name, review_title, review_content, stars, version, store]
        ws.append(review_final_data)
    print(f'{appid}: Reviews database created successfully!')
    wb.save('AppData.xlsx')
    wb.close()


def main():
    appid = get_play_id()
    for app in appid:
        details(app)
        review_all(app)



if __name__ == '__main__':
    main()



